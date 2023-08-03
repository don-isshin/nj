import codecs
import xlrd
import xlsxwriter
from lxml import etree
import openpyxl
import os
import shutil
import time
from openpyxl.styles import Font
import hashlib
import pandas as pd
from re import match
import generalTemplate

# 运行脚本日期
ndate = time.strftime('%Y%m%d', time.localtime(time.time()))
ndate1 = time.strftime('%Y-%m-%dT%H:%M:%S', time.localtime(time.time()))
# ndate2 = time.strftime('%Y%m%d', time.localtime(time.time()))

KJ_root = 'F:\\ShareFolderForLinux\\KUOJINGtest'
delivery_path = 'F:\\KUOJINGtest\\kjdelivery' + ndate + '\\'
rootdir = 'F:\\KUOJINGtest\\'
summary_path = 'F:\\KUOJINGtest\\summary'
# ndate = time.strftime('%Y%m%d', time.localtime(time.time()))
# ndate1 = time.strftime('%Y-%m-%dT%H:%M:%S', time.localtime(time.time()))
# ndate2 = time.strftime('%Y%m%d', time.localtime(time.time()))
datavalue = []
dire = []
dire[:] = []
dire2 = []
dire2[:] = []

try:
    os.mkdir(delivery_path)  # 不存在创建
except:
    pass

# 计算文件内容md5值
def md5sum(filename):
    file_object = open(filename, 'rb')  # 这段代码定义了一个名为`md5sum`的函数，该函数接受一个文件名作为参数，计算该文件的MD5值并返回。具体实现过程是通过二进制方式打开文件，读取文件内容，计算文件内容的MD5值，最后返回该值。
    file_content = file_object.read()
    file_object.close()
    file_md5 = hashlib.md5(file_content)
    return file_md5



# 打开一个excel文件
def open_xls(file):
    fh = xlrd.open_workbook(file)
    return fh


# 获取excel中所有的sheet表
def getsheet(fh):
    return fh.sheets()


# 获取sheet表的行数
def getnrows(fh, sheet):
    table = fh.sheets()[sheet]
    return table.nrows


# 读取文件内容并返回行内容
def getFilect(file, shnum):
    fh = open_xls(file)
    table = fh.sheets()[shnum]
    num = table.nrows
    for row in range(1, num):
        rdata = table.row_values(row)
        datavalue.append(rdata)
    return datavalue


# 读取第一个文件内容并返回行内容
def getFilect1(file, shnum):
    fh = open_xls(file)
    table = fh.sheets()[shnum]
    num = table.nrows
    for row in range(num):
        rdata = table.row_values(row)
        datavalue.append(rdata)
    return datavalue


# 获取sheet表的个数
def getshnum(fh):
    x = 0
    sh = getsheet(fh)
    for sheet in sh:
        x += 1
    return x


# path下的所有目录，排除 '.' 开头的文件
def listdir_nohidden(path):
    for f in os.listdir(path):
        if not f.startswith('.'):
            yield f

dire = []
dire[:] = []

dire2 = []
dire2[:] = []

for i in listdir_nohidden(KJ_root):
    dire.append(i)

for j in listdir_nohidden(KJ_root):
    for i in listdir_nohidden(KJ_root + '\\' + j):
        # print(i + '------------------------------------------------------------------')
        dire2.append(i)
datavalue = []


# 读取xml生成专辑excel和曲目excel，返回文件md5值
# rpath 是 xml 绝对路径，rrpath 是关于交付时间的目录的子目录
            # 该段代码用于生成一个XML文件，并将其写入指定路径下的文件中。
            # 如果指定路径下已存在同名文件，则会先删除该文件，再写入新的XML文件。
            # XML文件的内容包括：MessageHeader、AckonwledgedFile和FileStatus等元素。
            # 如果执行成功，则会输出“执行成功发送短信”。
            # 如果执行失败，则会输出DisplayTitle[0].text。
            # 最后，该段代码还会生成一个Excel文件，并将其写入指定路径下的文件中。
            # Excel文件的内容包括：ICPN、DeliveryTime、releaseTpye、displayartistName、formalTitle、DisplayTitle、genre、pline、ReleaseDate和audition等元素。
def xmlanalyzer(rpath, rrpath):
    print(rpath + "   " + rrpath)
    tree = etree.parse(rpath)
    root = tree.getroot()
    # cat_num = tree.findall('//ReleaseList/Release/ReleaseId/ICPN')
    icpn = tree.findall('//ReleaseList/Release/ReleaseId/ICPN')
    messagetype = tree.findall('//UpdateIndicator')
    release = tree.findall('//ReleaseList/Release/ReleaseType')
    dis_artists = tree.findall('//ReleaseList/Release/ReleaseDetailsByTerritory/DisplayArtistName')
    FormalTitle = tree.findall(
        '//ReleaseList/Release/ReleaseDetailsByTerritory/Title[@TitleType="FormalTitle"]/TitleText')
    DisplayTitle = tree.findall(
        '//ReleaseList/Release/ReleaseDetailsByTerritory/Title[@TitleType="DisplayTitle"]/TitleText')
    genre = tree.findall('//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/Genre/GenreText')
    pline = tree.findall('//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/PLine/PLineText')
    #pline = tree.findall('//ReleaseList/Release/PLine/PLineText')
    #/ernm:NewReleaseMessage/ReleaseList/Release[1]/PLine/PLineText
    # releasedate = tree.findall('//ReleaseList/Release/ReleaseDetailsByTerritory/ReleaseDate')
    releasedate = tree.findall('//DealList/ReleaseDeal/Deal/DealTerms/ValidityPeriod/StartDate')
    filenames = tree.findall(
        '//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/TechnicalSoundRecordingDetails/File/FileName')
    img = tree.findall('//ResourceList/Image/ImageDetailsByTerritory/TechnicalImageDetails/File/FileName')
    #/ernm:NewReleaseMessage/ResourceList/Image/ImageDetailsByTerritory/TechnicalImageDetails/File/FileName
    checksum = tree.findall('//HashSum/HashSum')
    deliverytime = ndate
    dtime = ndate + '000000001'
    takedown = tree.findall('DealList/ReleaseDeal/Deal/DealTerms/TakeDown')
    messagesenderid = tree.findall('//MessageHeader/MessageRecipient/PartyId')
    # 解析XML文件中的消息发送者ID。通过`tree.findall()`方法找到XML文件中所有`<MessageHeader><SentOnBehalfOf><PartyId>`标签，
    # 然后将它们保存在一个列表`messagesenderid`中。这个列表中的每个元素都是一个包含了标签内容的字符串。
    messagesendername = tree.findall('//MessageHeader/MessageRecipient/PartyName/FullName')
    MessageRecipientid = tree.findall('//MessageHeader/MessageSender/PartyId')
    MessageRecipientname = tree.findall('//MessageHeader/MessageSender/PartyName/FullName')
    MessageCreatedDateTime = ndate
    ReleaseId = tree.findall('//MessageHeader/MessageId')
    description = tree.findall('//ReleaseList/Release/ReleaseDetailsByTerritory/Synopsis')

    # 打开xlsx
    wb = openpyxl.Workbook()
    sheet = wb.active
    audition_path = 'http://audioaudition.sonyselect.cn/KUOJINGtest/delivery/' + rrpath
    tdre = []
    if messagetype[0].text != 'OriginalMessage':
        if takedown:
            tdalbum = 'ICPN: ' + icpn[0].text + '; The resources: '
            print('takedown')
            for i in takedown:
                tdresouce = i.xpath('../../../DealReleaseReference')
                tdre.append(tdresouce[0].text)               
                strdiv = ','
            tdre = set(tdre)
            td = tdalbum + strdiv.join(tdre)
            

            #print(td + '---------------------------------------------------------')
            with codecs.open('F:\\KUOJINGtest\\takedown' + '\\takedown.log', 'a', encoding='utf-8') as file2:
                file2.writelines("\r\n" + td)
            file2.close()
            ackfile = KJ_root  + '\\' + 'logs\\' + 'ACK_' + icpn[0].text + '_takedown' + '.xml'
            if os.path.exists(ackfile):
                os.remove(ackfile)

            with codecs.open(ackfile, 'w', encoding='utf-8') as file1:
                file1.writelines('<?xml version="1.0" encoding="utf-8"?><ns3:FtpAcknowledgementMessage xmlns:ns2="http://www.w3.org/2000.09/xmldsig#" xmlns:ns3="http://ddex.net/xml/ern-c/14" MessageVersionId="1.0"><MessageHeader><MessageSender><PartyId>' + 
                    messagesenderid[0].text + '</PartyId><PartyName><FullName>' + messagesendername[0].text + '</FullName></PartyName></MessageSender><MessageRecipient><PartyId>' + 
                    MessageRecipientid[0].text + '</PartyId><PartyName><FullName>' + MessageRecipientname[0].text +'</FullName></PartyName></MessageRecipient><MessageCreatedDateTime>' + 
                    ndate + '</MessageCreatedDateTime></MessageHeader><AckonwledgedFile><ReleaseId>' + 
                    ReleaseId[0].text + '</ReleaseId><ReleaseReference>' + td + '</ReleaseReference><Date>' + ndate + '</Date></AckonwledgedFile><FileStatus>takedown</FileStatus></ns3:FtpAcknowledgementMessage>')
                file1.close()
            # 执行成功呢发送短信
            # command = r'C:\Users\Administrator\Downloads\curl.exe -i -X "POST" -H "Content-Type: application/json" -H "Authorization: Basic RGF2aWQuQS5IYW5Ac29ueS5jb206V2VicDB3ZXI=" -d "{\"mobile\":\"18810259859\",\"content\":' + "\\\"%s\\\", \\\"campaignID\\\":966}\" https://sms.smsvip.cn/rest/v2/single_sms" % td
            # os.system(command)
            print("执行成功发送短信")
            # 写一个takedownack
            time.sleep(1)

            with codecs.open(delivery_path + '\\takedown.log', 'a', encoding='utf-8') as file7:
                file7.writelines("\r\n" + td)
            file7.close()
        else:
            ackfile = KJ_root  + '\\' + 'logs\\' + 'ACK_' + icpn[0].text + '.xml'
            if os.path.exists(ackfile):
                os.unlink(ackfile)
            with codecs.open(ackfile, 'w', encoding='utf-8') as file1:
                file1.writelines('<?xml version="1.0" encoding="utf-8"?><ns3:FtpAcknowledgementMessage xmlns:ns2="http://www.w3.org/2000.09/xmldsig#" xmlns:ns3="http://ddex.net/xml/ern-c/14" MessageVersionId="1.0"><MessageHeader><MessageSender><PartyId>' + 
                messagesenderid[0].text + '</PartyId><PartyName><FullName>' + messagesendername[0].text + '</FullName></PartyName></MessageSender><MessageRecipient><PartyId>' + 
                MessageRecipientid[0].text + '</PartyId><PartyName><FullName>' + MessageRecipientname[0].text +'</FullName></PartyName></MessageRecipient><MessageCreatedDateTime>' + 
                ndate + '</MessageCreatedDateTime></MessageHeader><AckonwledgedFile><ReleaseId>' + 
                ReleaseId[0].text + '</ReleaseId><Date>' + ndate +'</Date></AckonwledgedFile><FileStatus>File Exception</FileStatus></ns3:FtpAcknowledgementMessage>')
            file1.close()
    else:
       
        ackfile = KJ_root  + '\\' + 'logs\\' + 'ACK_' + icpn[0].text + '.xml'
        if os.path.exists(ackfile):
            os.unlink(ackfile)
        with codecs.open(ackfile, 'w', encoding='utf-8') as file1:
            file1.writelines('<?xml version="1.0" encoding="utf-8"?><ns3:FtpAcknowledgementMessage xmlns:ns2="http://www.w3.org/2000.09/xmldsig#" xmlns:ns3="http://ddex.net/xml/ern-c/14" MessageVersionId="1.0"><MessageHeader><MessageSender><PartyId>' + 
            messagesenderid[0].text + '</PartyId><PartyName><FullName>' + messagesendername[0].text + '</FullName></PartyName></MessageSender><MessageRecipient><PartyId>' + 
            MessageRecipientid[0].text + '</PartyId><PartyName><FullName>' + MessageRecipientname[0].text +'</FullName></PartyName></MessageRecipient><MessageCreatedDateTime>' + 
            ndate + '</MessageCreatedDateTime></MessageHeader><AckonwledgedFile><ReleaseId>' + 
            ReleaseId[0].text + '</ReleaseId><Date>' + ndate +'</Date></AckonwledgedFile><FileStatus>File Exception</FileStatus></ns3:FtpAcknowledgementMessage>')
        file1.close()
        print(DisplayTitle[0].text)
        with codecs.open(delivery_path + rrpath + '\\album_description.txt', 'a', encoding='utf-8') as filedes:
                filedes.writelines("\r\n" + description[0].text)
        value = [['ICPN', 'DeliveryTime', 'releaseTpye', 'displayartistName', 'formalTitle', 'displayTitle', 'genre',
                  'pline', 'ReleaseDate', 'audition'],
                 [icpn[0].text, dtime, release[0].text, dis_artists[0].text, FormalTitle[0].text, DisplayTitle[0].text,
                  genre[0].text, pline[0].text, releasedate[0].text, audition_path]]
        for j in range(0, len(value[0])):
            cf = sheet.cell(row=1, column=j + 1, value=str(value[0][j]))
            cf.font = Font(bold=True)
            sheet.cell(row=2, column=j + 1, value=str(value[1][j]))
        fileexcel = delivery_path + rrpath + '\\' + icpn[0].text + '_' + deliverytime + '_album.xlsx'
        print(delivery_path + rrpath + '\\' + icpn[0].text + '_' + deliverytime + '_album.xlsx' + ' generated')
        wb.save(fileexcel)
        wb.close()

    #  below is tracks info in album.

    ticpn = tree.findall('//ReleaseList/Release/ReleaseId/ICPN')
    srisrc = tree.findall('ResourceList/SoundRecording/SoundRecordingId/ISRC')
    tisrc = tree.findall('ReleaseList/Release/ReleaseId/ISRC')
    # tduration = tree.findall('//ResourceList/SoundRecording/Duration')
    # tdur = []
    # tdur[:] = []
    # for i in tduration:

        # ti = i.text
        # if 'M' not in ti:
            # s = ti.split('S')[0]
            # second = s.split('PT')[1]
            # sec = int(second)
            # tt = sec
            # t = time.strftime('%H:%M:%S', time.gmtime(tt))
            # tdur.append(t)
        # elif 'H' not in ti:
            # s = ti.split('S')[0]
            # second = s.split('M')[1]
            # s2 = s.split('M')[0]
            # minute = s2.split('PT')[1]
            # sec = int(second)
            # min = int(minute)
            # tt = min * 60 + sec
            # t = time.strftime('%H:%M:%S', time.gmtime(tt))
            # tdur.append(t)
        # else:
            # s = ti.split('S')[0]
            # second = s.split('M')[1]
            # s2 = s.split('M')[0]
            # minute = s2.split('H')[1]
            # s3 = s2.split('H')[0]
            # hour = s3.split('PT')[1]
            # sec = int(second)
            # min = int(minute)
            # hour = int(hour)
            # tt = hour * 60 * 60 + min * 60 + sec
            # t = time.strftime('%H:%M:%S', time.gmtime(tt))
            # tdur.append(t)
    tDisplayTitle = tree.findall(
        '//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/Title[@TitleType="DisplayTitle"]/TitleText')

    tFormalTitle = tree.findall('//ResourceList/SoundRecording/ReferenceTitle/TitleText')

    tRelType = []
    for i in tisrc:
        print(i)
        tReleaseType = i.xpath('../../ReleaseType')
        #/ernm:NewReleaseMessage/ReleaseList/Release[1]/ReleaseType
        #/ernm:NewReleaseMessage/ResourceList/SoundRecording[1]/SoundRecordingId/ISRC
        #/ernm:NewReleaseMessage/ReleaseList/Release[2]/ReleaseId/ISRC
        #/ernm:NewReleaseMessage/ReleaseList/Release[1]/ReleaseDetailsByTerritory/DisplayArtistName
        for j in tReleaseType:
            tRelType.append(j.text)
            print(j.text)
    tdisArtName = []
    for i in tisrc:
        tdisplayArtistName = i.xpath('../../ReleaseDetailsByTerritory/DisplayArtistName')
        for j in tdisplayArtistName:
            tdisArtName.append(j.text)
    tresource = []
    for i in srisrc:
        tresourcecon = i.xpath('../../SoundRecordingDetailsByTerritory/ResourceContributor/PartyName/FullName')
        if tresourcecon != []:
            for j in tresourcecon:
                tresource.append(j.text)
        else:
            tresource.append('')

    tdisplayArtist = []
    tdisArtist = []
    for i in tisrc:
        tdisplayArtist = i.xpath('../../ReleaseDetailsByTerritory/DisplayArtist/ArtistRole')
        for j in tdisplayArtist:
            if j.text == 'MainArtist':
                tdisplay = j.xpath('../PartyName/FullName')
                for k in tdisplay:
                    tdisArtist.append(i.text + '|' + k.text)
    trackdisartist = []
    sum = ''
    for i in tisrc:
        for j in tdisArtist:
            if j.split('|')[0] == i.text:
                singlear = j.split('|')[1]
                sum = sum + singlear + ';'
        trackdisartist.append(sum.strip(';'))
        sum = ''

    tresourceref = tree.xpath('//ResourceList/SoundRecording/ResourceReference')
    tdiscn = []
    tdiscp = tree.xpath(
        '//ReleaseList/Release/ReleaseResourceReferenceList/ReleaseResourceReference[@ReleaseResourceType="PrimaryResource"]')

    tgenre = tree.findall('//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/Genre/GenreText')

    tpline = tree.findall('//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/PLine/PLineText')
  

    taudiotype = tree.findall(
        '//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/TechnicalSoundRecordingDetails/AudioCodecType')
        #/ernm:NewReleaseMessage/ResourceList/SoundRecording[1]/SoundRecordingDetailsByTerritory/TechnicalSoundRecordingDetails/AudioCodecType

    tsamplingRate = tree.findall(
        '//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/TechnicalSoundRecordingDetails/SamplingRate')

    tmd5 = tree.findall(
        '//ResourceList/SoundRecording/SoundRecordingDetailsByTerritory/TechnicalSoundRecordingDetails/File/HashSum/HashSum')

    md5file = []
    ppath = rpath.split('\\')[:-1]
    newpath = '\\'.join(ppath)

    for i in checksum:
        tmd5file = i.xpath('../../FileName')
        for j in tmd5file:
            md5file.append(newpath + '\\resources\\' + j.text)

    wb = openpyxl.Workbook()
    sheet = wb.active
    albuminfotitle = sheet.cell(row=1, column=1, value='AlbumInfo')
    albuminfotitle.font = Font(bold=True, size=12)
    if messagetype[0].text != 'OriginalMessage':
        pass
    else:
        value1 = [['ICPN', 'DeliveryTime', 'releaseTpye', 'displayartistName', 'formalTitle', 'displayTitle', 'genre',
                   'pline', 'ReleaseDate'],
                  [icpn[0].text, deliverytime, release[0].text, dis_artists[0].text, FormalTitle[0].text,
                   DisplayTitle[0].text, genre[0].text, pline[0].text, releasedate[0].text]]

        for j in range(0, len(value1[0])):
            c = sheet.cell(row=2, column=j + 1, value=str(value1[0][j]))
            c.font = Font(bold=True)
            sheet.cell(row=3, column=j + 1, value=str(value1[1][j]))

    resinfotitle = sheet.cell(row=5, column=1, value='ResourceInfo')
    resinfotitle.font = Font(bold=True, size=12)
    # 这段代码首先检查消息类型是否为“OriginalMessage”，如果不是，则跳过。否则，它将创建一个名为“value2”的列表，其中包含一个包含各种元素的子列表。然后，它使用循环将这些元素写入Excel工作表的第6行。接下来，它使用另一个循环将来自XML文件的数据写入Excel工作表的第7行及以下行。最后，它将Excel工作簿保存到磁盘并返回一些值，包括校验和、MD5哈希、消息发送方和接收方的ID和名称、发布ID和ICPN。
    if messagetype[0].text != 'OriginalMessage':
        pass
    else:
        value2 = [
            ['ICPN', 'ISRC', 'displayTitle', 'FormalTitle', 'subTitle', 'releaseType', 'displayArtistName',
             'genre', 'pline', 'displayArtist', 'resourceContributor', 'audioCodec', 'samplingRate', 'md5'], []]
        for k in range(0, len(value2[0])):
            d = sheet.cell(row=6, column=k + 1, value=str(value2[0][k]))
            d.font = Font(bold=True)

        for l in range(0, len(value2[0])):
            for i in range(0, len(tisrc)):
                count = i + 1
                print(tpline[i].text)
                cell = [ticpn[0].text, tisrc[i].text, tDisplayTitle[i].text, tFormalTitle[i].text, '',
                        tRelType[i], tdisArtName[i], tgenre[i].text, tpline[i].text, trackdisartist[i], tresource[i],
                        taudiotype[i].text, tsamplingRate[i].text, tmd5[i].text]

                sheet.cell(row=6 + count, column=l + 1, value=str(cell[l]))

        fileexcel = delivery_path + rrpath + '\\' + icpn[0].text + '.xlsx'
        wb.save(fileexcel)
        wb.close()
    return checksum, md5file, messagesenderid[0].text, messagesendername[0].text ,MessageRecipientid[0].text, MessageRecipientname[0].text, ReleaseId[0].text, icpn[0].text

# 合并
def merge_xlsx():
    # 定义要合并的excel文件列表
    allxls = []
    allxls[:] = []

    for parent, dirnames, filenames in os.walk(rootdir + 'summary\\'):  # 遍历summary 目录、子目录、文件
        for filename in filenames:
            allxls.append(parent + filename)  # 获取地址
    # 存储所有读取的结果
    print("allxls:")
    print(allxls)

    for fl in allxls:
        # if fl == allxls[0]:
        #        print(allxls[0])
        fh = open_xls(fl)
        x = getshnum(fh)
        for shnum in range(1):
            print("正在读取文件：" + str(fl) + "的第" + str(shnum) + "个sheet表的内容...")
            if fl == allxls[0]:
                rvalue = getFilect1(fl, shnum)
                print(rvalue)
            else:
                rvalue = getFilect(fl, shnum)
                print(rvalue)

    # 定义最终合并后生成的新文件
    endfile = 'F:\\KUOJINGtest\\summary' + '\\GrandVista_summary.xlsx'
    endfilewithdate = 'F:\\KUOJINGtest\\summary' + '\\GrandVista_summary_' + ndate + '.xlsx'
    wb1 = xlsxwriter.Workbook(endfile)
    wb2 = xlsxwriter.Workbook(endfilewithdate)

    # 创建一个sheet工作对象
    ws = wb1.add_worksheet()
    for a in range(len(rvalue)):
        for b in range(len(rvalue[a])):
            c = rvalue[a][b]
            ws.write(a, b, c)
    wb1.close()
    ws = wb2.add_worksheet()
    for a in range(len(rvalue)):
        for b in range(len(rvalue[a])):
            c = rvalue[a][b]
            ws.write(a, b, c)
    wb2.close()
    print("文件合并完成,在" + endfile)
    print("文件合并完成,在" + endfilewithdate)

    df = pd.read_excel(endfilewithdate, dtype=str)
    df.sort_values(by=['ReleaseDate'], inplace=True, ascending=False)
    df.to_excel(endfilewithdate, index=False)
    writer = pd.ExcelWriter(endfilewithdate)

    #           目标路径     表名         不显示行索引   NaN？
    df.to_excel(writer, sheet_name='GrandVista', index=False, na_rep='NaN')
    for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets['GrandVista'].set_column(col_idx, col_idx, 20)

    writer.save()
    # input("Press <enter>")


###########
# 清空summary，复制到summary，处理xlsx
# 复制并重命名所有的 _album.xlsx 文件到 summary 目录下，并将文件名中的 '_album' 改为 '_sum'
def copy_summary():
    filelist = []
    for i in dire2:
        filelist.append(i + '_' + ndate + '_album.xlsx')
        print(filelist)

    for i in dire:
        filelist.append(i + '_' + ndate + '_album.xlsx')
        # filelist.append(i + '_album.xlsx')
        for parent, dirnames, filenames in os.walk(KJ_root + '\\' + i, followlinks=True):
            for filename1 in filenames:
                # print(filename1 + '+++++++++')
                if filename1.split('.')[-1] == 'xml' and filename1.split('_')[0] != 'ACK' and filename1.split('_')[
                    0] != 'BatchComplete':
                    # print(filename1)
                    xml_file_path = os.path.join(parent, filename1)
                    relpath = parent.split('\\')[-1]
                    print(delivery_path + relpath)
                    generalTemplate.generalXlsx(xml_file_path, delivery_path + relpath + '\\' + relpath.split('_')[
                        0] + 'publishingtemp.xlsx')

    # dir = rootdir + dir_input
    #
    ############
    # 如果存在xlsx文件，清空summary
    if os.path.exists(rootdir + 'summary\\' + 'GrandVista_summary.xlsx'):
        finddir = rootdir + 'summary\\'
        for parent, dirnames, filenames in os.walk(finddir):
            print("******parent:" + parent)
            for filename2 in filenames:
                # print("parent + filename2:" + parent + filename2)
                os.remove(parent + filename2)
                # print("清空summary")
    else:
        pass
        # print("清空summary2")

    print(filelist)

    for parent, dirnames, filenames in os.walk(rootdir):
        # parent = parent.replace('\\', '/')
        # print("----parent:" + parent)
        for filename in filenames:
            if filename[-11:] == '_album.xlsx':  # 以 "_album.xlsx" 结尾
                # print(filename)   # 4895241400896_20230223_album.xlsx
                # print(filename[-11:]) # _album.xlsx
                # if filename in filelist:
                fullpath = parent + '\\' + filename
                # print("--------------------------")
                # print(fullpath)     # F:/EVO/delivery\4895241400896/4895241400896_20230223_album.xlsx
                # print(fullpath[:16])    # F:/EVO/delivery\
                # print("--------------------------")

                if fullpath[:16] == 'F:\\KUOJINGtest\\delivery':
                    pass
                else:
                    print("fullpath: " + fullpath)

                    # 移动文件并改名
                    # 暂时注释
                    shutil.copy(fullpath, rootdir + 'summary\\' + filename.split('.')[0] + '_sum' + '.xlsx')

                    # print("fullpath:" + fullpath)
                    # rename_album_to_sum = filename.split('.')[0] + '_sum' + '.xlsx'
                    # print("rename_album_to_sum:" + rename_album_to_sum)
                    #
                    # summary_path = rootdir + 'summary/'
                    # print("summary_path:" + summary_path)
                    # shutil.copy(fullpath, rename_album_to_sum)
                    # 拷贝，改名 '_sum.xlsx'
            else:
                pass


######### move.py

# 将指定目录下的文件按照一定规则归档到指定目录下的 delivery/ 和 pdf/ 目录中
def move():
    dire = []
    dire[:] = []
    for i in listdir_nohidden(KJ_root):
        print("file:" + i)
        dire.append(i)

    delxmlist = []
    delxmlist[:] = []

    dirdate = []
    dirdate[:] = []

    direc = 'kjdelivery' + ndate
    # direc = 'idoldelivery' + '20210524'
    print("direc:" + direc)

    # 遍历指定目录下的所有目录、子目录和文件，包括软连接
    for parent, dirnames, filenames in os.walk(rootdir, followlinks=True):
        for dir in dirnames:
            # 将direc（一个目录名的列表）中的每个元素与当前遍历到的目录名进行匹配
            match_dir = match(direc, dir)
            if match_dir:
                dirdate.append(dir)  # 目录名匹配成功，则将该目录名加入dirdate列表中

    # 把pdf文件归纳到/pdf中

    # 目录名+文件名 拼接
    dire2 = []
    dire2[:] = []
    for i in listdir_nohidden(KJ_root):
        for j in listdir_nohidden(KJ_root + '\\' + i):
            print("dir2:" + j)
            if os.path.isdir(KJ_root + '\\' + i + '\\' + j):
                dire2.append(j)

    for j in dirdate:
        # 存在日志文件
        if os.path.exists(rootdir + j + '\\filelist.log'):
            for i in dire2:
                print(rootdir + j + '\\' + i)
                # 遍历 EVO / 包括软连接
                for parent, dirnames, filenames in os.walk(rootdir + j + '\\' + i, followlinks=True):
                    for filename1 in filenames:  # 遍历到的文件名
                        file_path = os.path.join(parent, filename1)  # filename1拼接目录，当前循环中遍历到的文件的完整路径
                        print(file_path)
                        if os.path.exists(file_path):  # 存在目录
                            if not os.path.exists(
                                    rootdir + 'delivery/' + file_path.split('\\')[-2]):  # 不存在EVOdelivery+文件路径
                                os.mkdir(rootdir + 'delivery/' + file_path.split('\\')[-2])  # 创建出来
                                shutil.copy(file_path, rootdir + 'delivery/' + file_path.split('\\')[-2] + '/' +
                                            file_path.split('\\')[-1])
                                if file_path.split('.')[-1] == 'pdf':
                                    # 拷贝 .pdf 到 pdf/ 路径下
                                    shutil.copy(file_path, rootdir + 'pdf/' + file_path.split('\\')[-1])
                            else:
                                shutil.copy(file_path, rootdir + 'delivery/' + file_path.split('\\')[-2] + '/' +
                                            file_path.split('\\')[-1])
                                if file_path.split('.')[-1] == 'pdf':
                                    shutil.copy(file_path, rootdir + 'pdf/' + file_path.split('\\')[-1])
                        else:
                            print('no such file path')
        else:
            print('no filelist.log')

######## ack.py

# 读取xml生成专辑excel和曲目excel，返回文件md5值
# rpath 是 xml 绝对路径，rrpath 是关于交付时间的目录的子目录
def xmlanalyzer_ack(rpath, rrpath, delpath):
    """
    读取XML文件并解析其中的内容，生成ACK文件并写入指定路径下的文件中。

    Args:
    rpath: str
        XML文件的绝对路径。
    rrpath: str
        关于交付时间的目录的子目录。
    delpath: str
        交付时间。

    Returns:
    None

    Raises:
    None
    """
    # rpath is xml abusolute path, rrpath is sub of directory about time of delivery
    tree = etree.parse(rpath)
    root = tree.getroot()
    print(rpath)
    icpn = tree.findall('//ReleaseList/Release/ReleaseId/ICPN')
    messagesenderid = tree.findall('//MessageHeader/MessageRecipient/PartyId')
    # 解析XML文件中的消息发送者ID。通过`tree.findall()`方法找到XML文件中所有`<MessageHeader><SentOnBehalfOf><PartyId>`标签，
    # 然后将它们保存在一个列表`messagesenderid`中。这个列表中的每个元素都是一个包含了标签内容的字符串。
    messagesendername = tree.findall('//MessageHeader/MessageRecipient/PartyName/FullName')
    MessageRecipientid = tree.findall('//MessageHeader/MessageSender/PartyId')
    MessageRecipientname = tree.findall('//MessageHeader/MessageSender/PartyName/FullName')
    MessageCreatedDateTime = ndate
    ReleaseId = tree.findall('//MessageHeader/MessageId')
    messagetype = tree.findall('//UpdateIndicator')

    print(ReleaseId[0].text + '----------------------')
    # 如果 EVO/交付时间目录 下不存在ack_log目录，则创建
    # if not os.path.exists(KJ_root + '\\' + rrpath + '\\' + 'logs\\'):
    #     os.mkdir(KJ_root + '\\' + rrpath + '\\' + 'logs\\')
    ackfile = KJ_root  + '\\' + 'logs\\' + 'ACK_' + icpn[0].text + '.xml'
    # ackfile = KJ_root + '\\' + rrpath + '\\' + 'logs\\' + 'ACK_' + delpath.split('_')[0] + '.xml'
    if os.path.exists(ackfile):  # 如果存在xml，移除
       os.unlink(ackfile)
    #if os.path.exists(ackfile) and messagetype[0].text != 'OriginalMessage':     
    # if messagetype[0].text != 'OriginalMessage':
    #     pass
    # if icpn[0].text in :
    #     pass
    #else:
    with codecs.open(ackfile, 'w', encoding='utf-8') as file1:
        file1.writelines('<?xml version="1.0" encoding="utf-8"?><ns3:FtpAcknowledgementMessage xmlns:ns2="http://www.w3.org/2000.09/xmldsig#" xmlns:ns3="http://ddex.net/xml/ern-c/14" MessageVersionId="1.0"><MessageHeader><MessageSender><PartyId>' + 
        messagesenderid[0].text + '</PartyId><PartyName><FullName>' + messagesendername[0].text + '</FullName></PartyName></MessageSender><MessageRecipient><PartyId>' + 
        MessageRecipientid[0].text + '</PartyId><PartyName><FullName>' + MessageRecipientname[0].text +'</FullName></PartyName></MessageRecipient><MessageCreatedDateTime>' + 
        ndate + '</MessageCreatedDateTime></MessageHeader><AckonwledgedFile><ReleaseId>' + 
        ReleaseId[0].text + '</ReleaseId><Date>' + ndate +'</Date></AckonwledgedFile><FileStatus>FileOK</FileStatus></ns3:FtpAcknowledgementMessage>')
    file1.close()
    return
    # print(KJ_root + '\\' + rrpath + '\\' + 'ack_log\\' + 'ACKtest_' + rrpath.split('_')[0] + '.xml')


def ack():
    # 创建空列表
    dire = []
    dire[:] = []
    for i in listdir_nohidden(KJ_root):
        dire.append(i)

    delxmlist = []
    delxmlist[:] = []
    for i in dire:
        # 目录、子目录、所有文件 遍历，包括软连接文件
        for parent, dirnames, filenames in os.walk(KJ_root + '\\' + i, followlinks=True):
            for filename1 in filenames:
                # 是xml文件，以ACK开头，  不是ACKtest开头，不是BatchComplete
                if filename1.split('.')[-1] == 'xml' and filename1.split('_')[0] != 'ACK' and filename1.split('_')[
                    0] != 'ACKtest' and filename1.split('_')[
                    0] != 'BatchComplete':
                    print(os.path.join(parent, filename1))
                    print(parent.split('\\')[-1])  # xml文件名所在目录
                    xml_file_path = os.path.join(parent, filename1)  # xml文件路径
                    relpath = i + '\\' + parent.split('\\')[-1]
                    delpath = parent.split('\\')[-1]  # 取xml文件名

                    # if not os.path.exists(delivery_path + relpath):
                    # os.mkdir(delivery_path + relpath)
                    xmla = xmlanalyzer_ack(rpath=xml_file_path, rrpath=relpath, delpath=delpath)
                    # `rpath`: XML文件路径，即要进行分析的XML文件的完整路径。
                    # `rrpath`: 相对路径，即与XML文件相对应的其他文件的路径。
                    # `delpath`: 删除路径，即要删除的XML节点的路径。
                    # print(relpath)
                    delxmlist.append(xml_file_path)
                    print('Deal with ' + xml_file_path + ' will be deleted.')
                else:
                    # print('not found xml file')
                    pass

    delxml = 'yes'
    if delxml.lower() == 'yes' or delxml.lower() == 'y':
        for i in delxmlist:
            # yield
            print(i)
            os.remove(i)

    # 删除bos文件
    # os.system('C:\\windows-bcecmd-0.3.0\\bcecmd bos sync --delete --yes F:\\ShareFolderForLinux\\360RA\\blank bos:/evog')
    # print("删除bos文件")
    deldir = []

    # 目录、子目录、所有文件
    for parent, dirnames, filenames in os.walk('F:\\ShareFolderForLinux\\KUOJINGtest', followlinks=True):
        for filename1 in filenames:
        #print(parent.split('\\')[2])
            if 'resources' in parent:
                print(parent.split('\\')[3])
                deldir.append(parent.split('\\')[3])  # 删除前缀，取第4级目录
        

    deldir = set(deldir)

    ############ 删除 deldir 目录
    print(deldir)
    for i in deldir:
        # 递归删除，忽略错误
        shutil.rmtree('F:\\ShareFolderForLinux\\KUOJINGtest\\' + i, ignore_errors=True)
        print("删除ftpKUOJINGtest列表中文件" + i)
        #os.system('C:/ClipMaker/ssh.exe.lnk root@ftp.sonyselect.com.cn "cd /opt/app/chroot/home/EVO; rm -rf %s"' % (i))


## renameaudition.py
# 读取文件内容并返回行内容
def getFilect_url(file, shnum):
    fh = open_xls(file)
    table = fh.sheets()[shnum]
    num = table.nrows
    for row in range(0, num):
        rdata = table.row_values(row)
        datavalue.append(rdata)

        if '180.76.140.158' in rdata[-3]:
            rdata[-3] = rdata[-3].replace('http://180.76.140.158', 'https://audioaudition.sonyselect.cn')

        if 'http://audioaudition.sonyselect.cn' in rdata[-3]:
            rdata[-3] = rdata[-3].replace('http://audioaudition.sonyselect.cn', 'https://audioaudition.sonyselect.cn')
        print(rdata[-3])
    return datavalue

def renameaudition():
    fl = 'F:\\KUOJINGtest\\summary\\GrandVista_summary_' + ndate + '.xlsx'

    datavalue = []

    fh = open_xls(fl)
    x = getshnum(fh)
    for shnum in range(1):
        print("正在读取文件：" + str(fl) + "的第" + str(shnum) + "个sheet表的内容...")
        rvalue = getFilect(fl, shnum)
        print(rvalue[-1])

    # 定义最终合并后生成的新文件

    endfilewithdate = 'F:\\KUOJINGtest\\summary' + '\\GrandVista_summary_auditon_' + ndate + '.xlsx'

    wb2 = xlsxwriter.Workbook(endfilewithdate)

    ws = wb2.add_worksheet()
    for a in range(len(rvalue)):
        for b in range(len(rvalue[a])):
            c = rvalue[a][b]
            ws.write(a, b, c)
    wb2.close()
    # print("文件合并完成,在" + endfile)
    print("文件合并完成,在" + endfilewithdate)
    df = pd.read_excel(endfilewithdate, dtype=str)
    df.sort_values(by=['ReleaseDate'], inplace=True, ascending=False)
    df.to_excel(endfilewithdate, index=False)
    writer = pd.ExcelWriter(endfilewithdate)

    df.to_excel(writer, sheet_name='GrandVista', index=False, na_rep='NaN')
    for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets['GrandVista'].set_column(col_idx, col_idx, 20)

    writer.save()


###############
# 删除不包含auditon的xlsx文件
def final_auditon():
    file_ext = ".xlsx"
    keyword = "auditon"
    for file in os.listdir(summary_path):
"""
This script performs various file operations such as deleting files, copying files, merging files, and executing batch files. 
It also generates logs for successful and failed operations. 
The script takes in a directory path and searches for XML files with specific naming conventions. 
It then performs MD5 checksum validation on the files and copies them to a delivery path. 
If the MD5 checksum validation fails, it generates an error log. 
The script also deletes files that do not contain a specific keyword and generates a file list log. 
Finally, it executes batch files to signal success or failure of the operation.
"""
        # 筛选出 .xlsx 文件，排除 auditon 关键字的文件
        if file.endswith(file_ext) and keyword not in file:
            file_path = os.path.join(summary_path, file)
            os.remove(file_path)
            print(f"已删除文件 {file_path}")


# 主程序入口
if __name__ == '__main__':
    try:
        # 遍历指定目录下的所有文件
        for i in dire:
            for parent, dirnames, filenames in os.walk(KJ_root + '\\' + i, followlinks=True):
                # 遍历目录下的所有文件
                for filename1 in filenames:
                    # 筛选出指定后缀名的文件
                    if filename1.split('.')[-1] == 'xml' and filename1.split('_')[0] != 'ACK' and filename1.split('_')[0] != 'BatchComplete':
                        ack_pass = False
                        xml_file_path = os.path.join(parent, filename1)
                        relpath = parent.split('\\')[-1]
                        # 如果目标目录不存在，则创建目标目录
                        if not os.path.exists(delivery_path + relpath):
                            os.mkdir(delivery_path + relpath)
                        # 对XML文件进行MD5校验
                        xml1 = xmlanalyzer(rpath=xml_file_path, rrpath=relpath)
                        # 将XML文件复制到目标目录
                        shutil.copy(xml_file_path, delivery_path + relpath + '\\')
                        shutil.copy(xml_file_path, 'F:\\KUOJINGtest\\kj_xml_backup\\')
                        # 遍历XML文件中的所有文件，进行MD5校验
                        for i, j in zip(xml1[0], xml1[1]):
                            file_md5 = md5sum(j)
                            rcheck = file_md5.hexdigest()
                            # 如果MD5校验通过，则将文件复制到目标目录，并记录到日志中
                            if i.text == rcheck:
                                shutil.copy(j, delivery_path + j.split('\\')[-3] + '\\')
                                with codecs.open(delivery_path + '\\finished.log', 'a', encoding='utf-8') as file2:
                                    file2.writelines("\r\n" + j + '    ' + rcheck + '  equal  ' + i.text)
                                file2.close()
                                with codecs.open(delivery_path + '\\filelist.log', 'a', encoding='utf-8') as file5:
                                    file5.writelines("\r\n" + j)
                                file5.close()
                            # 如果MD5校验失败，则记录到错误日志中
                            else:
                                with codecs.open(delivery_path + '\\error.log', 'a', encoding='utf-8') as file3:
                                    file3.writelines("\r\n" + j + '    ' + rcheck + '  not equal  ' + i.text)
                                file3.close()
                                failmd5 = 'no'
                                # ackfile = KJ_root  + '\\' + 'logs\\' + 'ACK_' + xml1[7] + '.xml'
                                # with codecs.open(ackfile, 'w', encoding='utf-8') as file1:
                                #        file1.writelines('<?xml version="1.0" encoding="utf-8"?><ns3:FtpAcknowledgementMessage xmlns:ns2="http://www.w3.org/2000.09/xmldsig#" xmlns:ns3="http://ddex.net/xml/ern-c/14" MessageVersionId="1.0"><MessageHeader><MessageSender><PartyId>' + 
                                #        xml1[2] + '</PartyId><PartyName><FullName>' + xml1[3] + '</FullName></PartyName></MessageSender><MessageRecipient><PartyId>' + 
                                #        xml1[4] + '</PartyId><PartyName><FullName>' + xml1[5] +'</FullName></PartyName></MessageRecipient><MessageCreatedDateTime>' + 
                                #        ndate + '</MessageCreatedDateTime></MessageHeader><AckonwledgedFile><ReleaseId>' + 
                                #        xml1[6] + '</ReleaseId><Date>' + ndate +'</Date></AckonwledgedFile><FileStatus>Files are not OK</FileStatus></ns3:FtpAcknowledgementMessage>')
                                #        file1.close()
                                # 如果MD5校验失败，则根据配置决定是否退出程序
                                if failmd5.lower() == 'yes' or failmd5.lower() == 'y':
                                    os._exit(0)
                                else:
                                   ack_pass = True
        # 根据配置决定是否删除文件
        delfile = 'yes'
        if delfile.lower() == 'yes' or delfile.lower() == 'y':
            with codecs.open(delivery_path + 'filelist.log', 'r', encoding='utf-8') as file4:
                try:
                    allfile = file4.readlines()
                    for i in allfile:
                        if i not in ['\n', '\r\n']:
                            # 如果文件存在，则删除文件
                            if os.path.exists(i.strip('\n\r')):
                                #os.unlink(i.strip('\n\r'))
                                print(i.strip('\n\r') + ' deleted!')
                            else:
                                print('Files not found!')
                finally:
                    file4.close()
        else:
            os._exit(0)

        print("执行combine.py")
        # stri = ('D:\devTools\Anaconda3\python.exe combine.py')
        # proc = os.system(stri)

        copy_summary()
        merge_xlsx()

        print("执行move.py")
        move()
        # stri = ('D:\devTools\Anaconda3\python.exe move.py')
        # proc = os.system(stri)

        print("执行ack.py")
        if ack_pass != True: 
            ack()
        # stri = ('D:\devTools\Anaconda3\python.exe ack.py')
        # proc = os.system(stri)

        print("执行success.bat")
        os.system("C:\\Windows\\System32\\cmd.exe /c F:\\KUOJINGtest\\service\\success.bat")
    except Exception as e:
        import traceback

        traceback.print_exc()
        # except:
        # print("执行fail.bat")
        os.system("C:\\Windows\\System32\\cmd.exe /c F:\KUOJINGtest\\service\\fail.bat")

    renameaudition()

    # 删除不包含 auditon 关键字的文件
    # final_auditon()

